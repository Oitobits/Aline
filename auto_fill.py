import os
import pypdf
import re
import openpyxl
from openpyxl.utils import get_column_letter

# Paths
BASE_DIR = r"C:\Users\Usuário\Desktop\Projetos\Edson"
PDF_PATH = os.path.join(BASE_DIR, "Laudo Completo 03_01_2026.pdf")
TEMPLATE_PATH = os.path.join(BASE_DIR, "Anamnese.xlsx")

def extract_pdf_data(pdf_path):
    print(f"Reading PDF: {pdf_path}")
    reader = pypdf.PdfReader(pdf_path)
    full_text = ""
    for page in reader.pages:
        full_text += page.extract_text() + "\n"

    def find_pattern(pattern, text, flags=re.IGNORECASE):
        match = re.search(pattern, text, flags)
        if match:
            return match.group(1).strip()
        return None

    # Patient info
    patient_name = find_pattern(r"([A-Za-z\s]+)\s+DN:\s+\d{2}/\d{2}/\d{4}", full_text)
    birth_date = find_pattern(r"[A-Za-z\s]+\s+DN:\s*(\d{2}/\d{2}/\d{4})", full_text)
    age = find_pattern(r"[A-Za-z\s]+\s+DN:\s*\d{2}/\d{2}/\d{4}\s*\((\d+\s*Anos)\)", full_text)
    collection_date = find_pattern(r"Data de Coleta/Recebimento:\s*(\d{2}/\d{2}/\d{4})", full_text)

    # Exam patterns
    patterns = {
        "Eritrócitos": r"Eritrócitos\s+([\d,.]+)",
        "Hemoglobina": r"Hemoglobina\s+([\d,.]+)",
        "Hematócrito": r"Hematócrito\s+([\d,.]+)",
        "VCM": r"VCM\s+([\d,.]+)",
        "HCM": r"HCM\s+([\d,.]+)",
        "CHCM": r"CHCM\s+([\d,.]+)",
        "RDW": r"RDW\s+([\d,.]+)",
        "Leucócitos": r"Leucócitos\s+\d+\s+([\d.]+)",
        "Neutrófilos": r"Neutrófilos\s+([\d,.]+)\s+[\d.]+",
        "Eosinófilos": r"Eosinófilos\s+([\d,.]+)\s+[\d.]+",
        "Basófilos": r"Basófilos\s+([\d,.]+)\s+[\d.]+",
        "Linfócitos": r"Linfócitos\s+([\d,.]+)\s+[\d.]+",
        "Monócitos": r"Monócitos\s+([\d,.]+)\s+[\d.]+",
        "Plaquetas": r"Contagem de Plaquetas\s+([\d.]+)",
        "VPM": r"VPM\s+([\d,.]+)",
        "Ferritina": r"Ferritina\s+([\d,.]+)",
        "Ácido Fólico": r"Ácido Fólico\s+(Superior a\s+[\d,.]+|[\d,.]+)",
        "Vitamina B-12": r"Vitamina B-12,\s*Dosagem\s*([\d,.]+)",
        "Ureia": r"Uréia\s+([\d,.]+)",
        "Creatinina": r"Creatinina\s+([\d,.]+)",
        "eGFR": r"\*eGFR\s+(Superior a\s+[\d,.]+|[\d,.]+)",
        "Potássio": r"Potássio\s+([\d,.]+)",
        "Sódio": r"Sódio\s+([\d,.]+)",
        "Cálcio": r"Cálcio\s+([\d,.]+)",
        "Vitamina D": r"25-Hidroxivitamina D\s*\(Vitamina D\)\s*([\d,.]+)",
        "Glicose": r"Glicose\s+([\d,.]+)\s+mg/dL",
        "HbA1c": r"Hemoglobina Glicada\s*-\s*HbA1c\s*([\d,.]+)\s*%",
        "GME": r"Glicose Média Estimada\s*\(GME\)\s*([\d,.]+)",
        "Colesterol Total": r"Colesterol Total\s+([\d,.]+)",
        "LDL": r"LDL - Colesterol\s*\(calculado\)\s*([\d,.]+)",
        "HDL": r"HDL - Colesterol\s+([\d,.]+)",
        "Triglicérides": r"Triglicérides\s+([\d,.]+)",
        "TGO": r"Transaminase oxalacética - TGO\s*\(Aspartato amino transferase\)\s*([\d,.]+)",
        "TGP": r"Transaminase pirúvica -\s*TGP\s*\(Alanina amino\s*transferase\)\s*([\d,.]+)",
        "GGT": r"Gama-Glutamil\s*Transferase\s*([\d,.]+)",
        "TSH": r"Hormônio\s*Tireoestimulante\s*Ultrassensível \(TSH\)\s*([\d,.]+)",
        "T4 Livre": r"Tiroxina Livre \(T4 Livre\)\s*([\d,.]+)",
        "Zinco": r"Zinco Sanguineo\s+([\d,.]+)"
    }

    results = {}
    for name, pattern in patterns.items():
        results[name] = find_pattern(pattern, full_text)

    return {
        "patient_name": patient_name,
        "birth_date": birth_date,
        "age": age,
        "collection_date": collection_date,
        "results": results
    }

def clean_value(param_name, val_str):
    if not val_str:
        return None
    val_str = val_str.strip()
    
    # Check if it's a string value like "Superior a 24,0" or "Não Reagente"
    if any(word in val_str.lower() for word in ["superior", "inferior", "não", "reagente"]):
        return val_str
        
    try:
        if param_name in ["Leucócitos", "Plaquetas"]:
            # Remove dots (thousands separator)
            clean_str = val_str.replace(".", "")
            return int(clean_str)
        else:
            # Replace comma with dot for float conversion
            clean_str = val_str.replace(",", ".")
            return float(clean_str)
    except ValueError:
        return val_str

def update_spreadsheet(template_path, data):
    print(f"Loading Spreadsheet Template: {template_path}")
    wb = openpyxl.load_workbook(template_path, data_only=False)
    
    # 1. Update general info sheet
    if 'Informações gerais' in wb.sheetnames:
        ws_info = wb['Informações gerais']
        ws_info['B3'] = data['patient_name']
        ws_info['B4'] = data['age']
        ws_info['B5'] = data['birth_date']
        print("Updated 'Informações gerais' (Patient details)")

    # 2. Update Plan sheet (P1)
    if 'P1' in wb.sheetnames:
        ws_p1 = wb['P1']
        ws_p1['O2'] = data['patient_name']
        ws_p1['T2'] = data['collection_date']
        print("Updated 'P1' (Patient name and exam date)")

    # 3. Update Exams sheet
    if 'Exames ' in wb.sheetnames:
        ws_ex = wb['Exames ']
        
        # Find parameter mapping
        param_mapping = {
            "Eritrócitos": "Hemácias ",
            "Hemoglobina": "Hemoglobina",
            "Hematócrito": "Hematócrito",
            "VCM": "VCM ",
            "HCM": "HCM ",
            "CHCM": "CHCM",
            "RDW": "RDW",
            "Leucócitos": "Leucócitos",
            "Neutrófilos": "Neutrófilo ",
            "Linfócitos": "Linfócitos ",
            "Eosinófilos": "Eosinófilos",
            "Monócitos": "Monócitos ",
            "Basófilos": "Basófilos",
            "Plaquetas": "Plaquetas",
            "VPM": "VPM ",
            "Glicose": "Glicemia jejum ",
            "HbA1c": "Hb1Ac",
            "Colesterol Total": "Colesterol Total ",
            "HDL": "HDL colesterol ",
            "Triglicérides": "Triglicérides ",
            "LDL": "LDL colesterol ",
            "TGO": "TGO/AST ",
            "TGP": "TGP/ALT ",
            "GGT": "GGT (gama GT)",
            "Creatinina": "Creatinina ",
            "Ureia": "Ureia sérica ",
            "TSH": "TSH ULTRA SENSÍVEL ",
            "T4 Livre": "T4 livre ",
            "Ferritina": "Ferritina ",
            "Vitamina B-12": "Vitamina B12 ",
            "Ácido Fólico": "Vitamina B9",
            "Zinco": "Zinco sérico ",
            "Vitamina D": "25 (OH) D ",
            "Sódio": "Sódio ",
            "Potássio": "Potassio ",
            "Cálcio": "Calcio Iônico ."
        }

        # Find the first available data column
        # Columns start at C (column 3). We check row 2 (which holds the date or "DATA")
        target_col = 3
        while True:
            cell_val = ws_ex.cell(row=2, column=target_col).value
            # If the cell is empty or is the placeholder "DATA", we can use this column
            if cell_val is None or str(cell_val).strip().upper() == "DATA":
                break
            target_col += 1
        
        col_letter = get_column_letter(target_col)
        print(f"Target Column for exam data: {col_letter} (Column index {target_col})")
        
        # Write the collection date in row 2 of target column
        ws_ex.cell(row=2, column=target_col, value=data['collection_date'])
        
        # Find row index for each parameter in column 1
        row_map = {}
        for r in range(1, ws_ex.max_row + 1):
            val = ws_ex.cell(row=r, column=1).value
            if val:
                row_map[str(val).strip().lower()] = r

        # Write results
        for pdf_name, val_str in data['results'].items():
            sheet_name = param_mapping.get(pdf_name)
            if sheet_name:
                key = sheet_name.strip().lower()
                if key in row_map:
                    row_idx = row_map[key]
                    cleaned_val = clean_value(pdf_name, val_str)
                    ws_ex.cell(row=row_idx, column=target_col, value=cleaned_val)
                    # print(f"Wrote {pdf_name} -> Row {row_idx}: {cleaned_val}")
                else:
                    print(f"Warning: Sheet parameter '{sheet_name}' not found in Exames sheet rows.")
            else:
                print(f"Warning: PDF parameter '{pdf_name}' not mapped to spreadsheet.")

        print(f"Updated 'Exames' sheet with results under date {data['collection_date']}")

    # Save to a new file named after the patient
    safe_name = "".join([c for c in data['patient_name'] if c.isalpha() or c==' ']).rstrip().replace(" ", "_")
    output_filename = f"Anamnese_{safe_name}.xlsx"
    output_path = os.path.join(BASE_DIR, output_filename)
    wb.save(output_path)
    print(f"Saved completed spreadsheet to: {output_path}")
    return output_path

if __name__ == "__main__":
    if not os.path.exists(PDF_PATH):
        print(f"Error: PDF file not found at {PDF_PATH}")
        sys.exit(1)
    if not os.path.exists(TEMPLATE_PATH):
        print(f"Error: Spreadsheet template not found at {TEMPLATE_PATH}")
        sys.exit(1)
        
    data = extract_pdf_data(PDF_PATH)
    out_path = update_spreadsheet(TEMPLATE_PATH, data)
    print("Success!")
