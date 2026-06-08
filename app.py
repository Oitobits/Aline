import os
import re
import sys
import uuid
import time
import pypdf
import openpyxl
import unicodedata
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from flask import Flask, request, jsonify, render_template, send_from_directory
from google import genai
from google.genai import types
from pydantic import BaseModel, Field
from typing import List, Optional

# Reconfigure stdout to use UTF-8 just in case
sys.stdout.reconfigure(encoding='utf-8')

app = Flask(__name__)

# Config
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'temp_uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024 # 16MB max

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Pydantic Schemas for Gemini Structured Output
class ExamResult(BaseModel):
    parameter: str = Field(description="Nome do parâmetro correspondente da planilha Excel, exatamente como fornecido na lista de parâmetros permitidos.")
    value: str = Field(description="O valor numérico bruto extraído como string (ex: '4,89', '14,8', '6.650', '230.000', 'Superior a 24,0', '102,8')")

class LabReport(BaseModel):
    patient_name: str = Field(description="Nome completo do paciente")
    birth_date: Optional[str] = Field(description="Data de nascimento do paciente (DD/MM/AAAA)")
    age: Optional[str] = Field(description="Idade do paciente (ex: '36 Anos')")
    collection_date: str = Field(description="Data de coleta ou recebimento do exame (DD/MM/AAAA)")
    results: List[ExamResult]

# Helper to clean/parse values
def clean_value(param_name, val_str):
    if not val_str:
        return None
    val_str = val_str.strip()
    
    # Check if it's text-based like "Superior a 24,0" or "Não Reagente"
    if any(word in val_str.lower() for word in ["superior", "inferior", "não", "reagente", "indetectável"]):
        return val_str
        
    try:
        # Check parameter type for thousands vs decimal separator
        lower_name = param_name.lower()
        if "leucócitos" in lower_name or "plaquetas" in lower_name or "leucocitos" in lower_name:
            # Remove dots (thousands separator in PT-BR)
            clean_str = val_str.replace(".", "")
            return int(clean_str)
        else:
            # Replace comma with dot for float conversion
            clean_str = val_str.replace(",", ".")
            # In case it has a thousands dot (e.g. 1.200,00)
            if clean_str.count('.') > 1:
                # remove all dots except the last one
                parts = clean_str.split('.')
                clean_str = "".join(parts[:-1]) + "." + parts[-1]
            return float(clean_str)
    except ValueError:
        return val_str

# Helper to normalize strings for comparison (lowercase, accents, spaces removed)
def normalize_str(s):
    if not s:
        return ""
    s = " ".join(s.strip().lower().split())
    return "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

# Helper to cleanup old files from upload directory (older than 15 minutes)
def cleanup_old_files():
    try:
        now = time.time()
        for filename in os.listdir(UPLOAD_FOLDER):
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            # 15 minutes = 900 seconds
            if os.path.isfile(file_path) and (now - os.path.getmtime(file_path)) > 900:
                try:
                    os.remove(file_path)
                except Exception as e:
                    print(f"Error deleting temp file {file_path}: {e}")
    except Exception as e:
        print(f"Error listing UPLOAD_FOLDER for cleanup: {e}")

@app.route('/')
def index():
    has_server_key = bool(os.environ.get('GEMINI_API_KEY'))
    return render_template('index.html', has_server_key=has_server_key)

@app.route('/process', methods=['POST'])
def process():
    # Trigger auto-cleanup of old files
    cleanup_old_files()
    
    api_key = request.form.get('api_key') or os.environ.get('GEMINI_API_KEY')
    if not api_key:
        return jsonify({'error': 'A Chave de API do Gemini é obrigatória.'}), 400
        
    if 'pdf' not in request.files or 'xlsx' not in request.files:
        return jsonify({'error': 'Ambos os arquivos PDF e Excel são obrigatórios.'}), 400
        
    pdf_file = request.files['pdf']
    xlsx_file = request.files['xlsx']
    
    if pdf_file.filename == '' or xlsx_file.filename == '':
        return jsonify({'error': 'Arquivos inválidos fornecidos.'}), 400

    # Save temp files
    session_id = str(uuid.uuid4())
    pdf_temp_name = f"{session_id}_exam.pdf"
    xlsx_temp_name = f"{session_id}_template.xlsx"
    
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_temp_name)
    xlsx_path = os.path.join(app.config['UPLOAD_FOLDER'], xlsx_temp_name)
    
    pdf_file.save(pdf_path)
    xlsx_file.save(xlsx_path)
    
    try:
        # 1. Extract text from PDF
        reader = pypdf.PdfReader(pdf_path)
        pdf_text = ""
        for page in reader.pages:
            pdf_text += page.extract_text() + "\n"
            
        if not pdf_text.strip():
            raise Exception("Não foi possível extrair texto do PDF. O laudo pode ser uma imagem escaneada.")

        # 2. Load Excel file and dynamically extract parameters for Gemini
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
        excel_parameters = []
        if 'Exames ' in wb.sheetnames:
            ws_ex = wb['Exames ']
            for r in range(3, ws_ex.max_row + 1):
                val = ws_ex.cell(row=r, column=1).value
                if val:
                    val_str = str(val).strip()
                    # Skip header labels/calculated formulas
                    if not any(indicator in val_str.lower() for indicator in ["divisão", "melhor preditor", "exames", "parâmetros", "exame de saliva", "cortisol salivar"]):
                        excel_parameters.append(val_str)

        # 3. Call Gemini Client with Structured Output
        client = genai.Client(api_key=api_key)
        
        # Format parameter list for the Gemini system instruction
        parameters_list_str = "\n".join([f"- {p}" for p in excel_parameters])
        
        system_instruction = (
            "Você é um especialista médico em transcrição de exames laboratoriais brasileiros.\n"
            "Sua tarefa é extrair com precisão do texto do laudo todos os resultados dos exames, o nome completo do paciente, "
            "sua idade, data de nascimento e a data de coleta dos exames.\n\n"
            "IMPORTANTE: Você deve mapear cada exame extraído para o parâmetro correspondente exato na seguinte lista de parâmetros da planilha:\n"
            f"{parameters_list_str}\n\n"
            "Regras de Mapeamento:\n"
            "1. Identifique sinônimos e abreviações comuns para fazer o mapeamento correto. Exemplos:\n"
            "   - 'Hemoglobina Glicada' ou 'A1c' deve ser mapeado para 'Hb1Ac'.\n"
            "   - 'Gama GT', 'GGT' ou 'Gama-Glutamil Transferase' deve ser mapeado para 'GGT (gama GT)'.\n"
            "   - 'Triglicerídeos' ou 'Triglicérides' deve ser mapeado para 'Triglicérides '.\n"
            "   - 'Cortisol' ou 'Cortisol Sérico' deve ser mapeado para 'Cortisol'.\n"
            "   - 'HOMA-IR' ou 'HOMA IR' deve ser mapeado para 'HOMA IR (RESISTÊNCIA A INSULINA)'.\n"
            "   - 'Saturação de Transferrina' deve ser mapeado para 'Saturação de transferrina  sérica  (jejum)'.\n"
            "   - 'Fosfatase Alcalina' deve ser mapeado para 'FA (fosfatase alcalina)'.\n"
            "   - 'Vitamina A' ou 'Retinol' deve ser mapeado para 'Retinol Vit A '.\n"
            "   - 'Calcio Ionico' ou 'Cálcio Ionizado' deve ser mapeado para 'Calcio Iônico .'.\n"
            "   - 'Cálcio Sérico' ou 'Cálcio Total' deve ser mapeado para 'Cálcio'.\n"
            "2. Retorne o nome do parâmetro EXATAMENTE como escrito na lista fornecida.\n"
            "3. Se um exame do laudo não tiver correspondente claro na lista de parâmetros da planilha, ignore-o."
        )

        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=pdf_text,
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                response_schema=LabReport,
                system_instruction=system_instruction
            ),
        )
        
        # Validate data
        extracted_data = LabReport.model_validate_json(response.text)
        
        # 4. Update Excel file
        # Update General Info
        if 'Informações gerais' in wb.sheetnames:
            ws_info = wb['Informações gerais']
            ws_info['B3'] = extracted_data.patient_name
            if extracted_data.age:
                ws_info['B4'] = extracted_data.age
            if extracted_data.birth_date:
                ws_info['B5'] = extracted_data.birth_date
                
        # Update Plan P1
        if 'P1' in wb.sheetnames:
            ws_p1 = wb['P1']
            ws_p1['O2'] = extracted_data.patient_name
            ws_p1['T2'] = extracted_data.collection_date
            
        # Update Exams
        if 'Exames ' in wb.sheetnames:
            ws_ex = wb['Exames ']
            
            # Find target column
            target_col = 3 # Column C is 3
            while True:
                cell_val = ws_ex.cell(row=2, column=target_col).value
                if cell_val is None or str(cell_val).strip().upper() == "DATA":
                    break
                target_col += 1
                
            # Write Date
            ws_ex.cell(row=2, column=target_col, value=extracted_data.collection_date)
            
            # Build parameter row mapping (exact lower and normalized)
            row_map = {}
            for r in range(1, ws_ex.max_row + 1):
                val = ws_ex.cell(row=r, column=1).value
                if val:
                    val_str = str(val)
                    row_map[val_str.strip().lower()] = r
                    row_map[normalize_str(val_str)] = r
                    
            # Fills for styling
            green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") # soft green
            
            # Categorize parameters for trend analysis
            lower_better = [
                "glicose", "glicemia jejum", "colesterol total", "triglicérides", "triglicerideos", 
                "ldl", "ldl colesterol", "tgo/ast", "tgp/alt", "ggt (gama gt)", "creatinina", 
                "ureia sérica", "hb1ac", "hemoglobina glicada", "insulina", "homa ir", "homa beta", 
                "pcr us", "homocisteína", "fibrinogênio", "t3 reverso", "anti-tg", "anti-tpo", 
                "trab", "tsi", "tireoglobulina", "ácido úrico", "fosfatase alcalina", "pth", 
                "anti-endomisio", "anti-transglutaminase", "ige total", "prolactina", 
                "lipoproteína a", "cea", "ca 125", "ca 19,9"
            ]
            higher_better = [
                "hemácias", "hemoglobina", "hematócrito", "hdl colesterol", "ferro sérico", 
                "ferritina", "vitamina b12", "vitamina b9", "zinco sérico", "25 (oh) d", 
                "vitamina d", "magnésio sérico", "selênio", "vitamina b6", "vitamina b1", 
                "vitamina c", "cromo serico", "testosterona livre", "testosterona total", 
                "progesterona", "estradiol", "estrona", "sdhea", "shbg", "fsh", "lh", "igf1"
            ]
            
            prev_col = target_col - 1
            
            # Write values and calculate trend comparison
            for result in extracted_data.results:
                extracted_param = result.parameter
                if not extracted_param:
                    continue
                    
                key = extracted_param.strip().lower()
                norm_key = normalize_str(extracted_param)
                
                row_idx = None
                if key in row_map:
                    row_idx = row_map[key]
                elif norm_key in row_map:
                    row_idx = row_map[norm_key]
                    
                if row_idx:
                    cleaned_val = clean_value(extracted_param, result.value)
                    
                    # Write value
                    ws_ex.cell(row=row_idx, column=target_col, value=cleaned_val)
                    
                    # Trend comparison (if previous column has value and both are numeric)
                    if prev_col >= 3:
                        prev_val = ws_ex.cell(row=row_idx, column=prev_col).value
                        if isinstance(cleaned_val, (int, float)) and isinstance(prev_val, (int, float)):
                            has_improved = False
                            
                            # Use row name for lookup to avoid mismatch with extracted name
                            row_name = str(ws_ex.cell(row=row_idx, column=1).value).lower()
                            
                            # Check logic
                            if any(lbl in row_name for lbl in lower_better):
                                if cleaned_val < prev_val:
                                    has_improved = True
                            elif any(hbl in row_name for hbl in higher_better):
                                if cleaned_val > prev_val:
                                    has_improved = True
                                    
                            if has_improved:
                                ws_ex.cell(row=row_idx, column=target_col).fill = green_fill
            
        # Save output to a unique filename
        safe_patient_name = "".join([c for c in extracted_data.patient_name if c.isalnum() or c==' ']).strip().replace(" ", "_")
        output_filename = f"Anamnese_{safe_patient_name}_{session_id[:6]}.xlsx"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        wb.save(output_path)
        
        # Cleanup input files
        if os.path.exists(pdf_path): os.remove(pdf_path)
        if os.path.exists(xlsx_path): os.remove(xlsx_path)
        
        # Return success with patient summary and download link
        return jsonify({
            'success': True,
            'patient_name': extracted_data.patient_name,
            'collection_date': extracted_data.collection_date,
            'age': extracted_data.age,
            'birth_date': extracted_data.birth_date,
            'filename': output_filename,
            'results': [r.model_dump() for r in extracted_data.results]
        })
        
    except Exception as e:
        # Cleanup
        if os.path.exists(pdf_path): os.remove(pdf_path)
        if os.path.exists(xlsx_path): os.remove(xlsx_path)
        print(f"Error processing: {e}")
        return jsonify({'error': f"Ocorreu um erro no processamento: {str(e)}"}), 500

@app.route('/download/<filename>')
def download(filename):
    # Prevent directory traversal attacks
    if ".." in filename or filename.startswith("/") or filename.startswith("\\"):
        return jsonify({'error': 'Acesso negado.'}), 403
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

if __name__ == '__main__':
    # Local dev server
    app.run(host='127.0.0.1', port=5000, debug=True)
